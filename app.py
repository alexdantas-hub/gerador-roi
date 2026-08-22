import re
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime, timedelta

st.set_page_config(page_title="ROI Intelligence System", layout="wide")

# Rotulo da linha que agrega toda a receita do AdX sem investimento no dia.
LINHA_RESIDUAL = "RECEITA RESIDUAL (SEM INVESTIMENTO)"


def color_negative_red(val):
    if isinstance(val, (int, float)):
        color = 'red' if val < 0 else 'green'
        return f'color: {color}; font-weight: bold'
    return ''


def get_gspread_client():
    scope = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    creds_dict = st.secrets["gcp_service_account"]
    creds = Credentials.from_service_account_info(creds_dict, scopes=scope)
    return gspread.authorize(creds)


def clean_numeric(val):
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace('R$', '').replace(' ', '').replace('%', '')
    if not s or s.lower() == 'nan':
        return 0.0
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except:
        return 0.0


def tem_numero(chave):
    """True se a chave gerada e um numero de campanha utilizavel (<= 6 digitos)."""
    return bool(re.match(r'^\d{1,6}$', str(chave)))


def gerar_chave(nome):
    """
    Chave robusta de pareamento Meta <-> AdX: o NUMERO da campanha.

    Padrao esperado dos nomes: <PREFIXO><NUMERO>.<NICHO>.<PAIS>.<DATA>
        AD162.BIENESTAR.MEX.13/08/2026
        ca162.BIENESTARQUIZ.MEX.13/08/2026

    Nicho, pais e data NAO entram na chave porque divergem entre as
    plataformas na pratica:
      - nicho:  bienestar  vs  bienestarquiz
      - nome truncado no Meta: 'AD164' (sem nicho/pais/data)
      - data URL-encodada no AdX: '192f082f2026' (%2F -> 2f)
    O numero da campanha ja e o identificador unico, entao a chave de
    'AD164' e de 'ad164.anses.arg.26/04/2026' vira '164' nos dois lados.

    IDs longos do Meta (> 6 digitos) sao preservados inteiros para que o
    mapa de IDs (id_para_core) resolva a linha.
    """
    s = str(nome).lower().strip().replace('"', '').replace('%2f', '/')
    s = re.sub(r'^[a-z]+[\W_]*', '', s)     # remove prefixo alfabetico (ad, ca, cad, gp_...)
    m = re.match(r'^(\d+)', s)
    if m and len(m.group(1)) <= 6:          # numero de campanha
        return m.group(1)
    return s                                # ID longo ou formato desconhecido


st.title("📊 ROI Intelligence System")

tab1, tab2 = st.tabs(["🚀 Processamento Diário", "📈 Consulta Histórica"])

with tab1:
    st.sidebar.header("Configurações")
    data_referencia = st.sidebar.date_input("Data de Referência", datetime.now())
    cambio = st.sidebar.number_input("Cotação do Dólar (R$)", value=5.00, step=0.01)

    # Inicializa chaves dos uploaders no session_state
    if 'uploader_key' not in st.session_state:
        st.session_state['uploader_key'] = 0

    col_u1, col_u2, col_u3 = st.columns([4, 4, 1])
    with col_u1:
        files_meta = st.file_uploader("📁 Meta Ads (uma ou mais contas)", type=["csv"],
                                      accept_multiple_files=True,
                                      key=f"meta_{st.session_state['uploader_key']}")
    with col_u2:
        file_adx = st.file_uploader("📁 AdX", type=["csv"],
                                    key=f"adx_{st.session_state['uploader_key']}")
    with col_u3:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🗑️ Limpar", help="Remove os arquivos carregados"):
            st.session_state['uploader_key'] += 1
            st.rerun()

    if files_meta and file_adx:
        try:
            # Consolida todos os arquivos do Meta em um único DataFrame.
            # "Nome da campanha" e a coluna de EXIBICAO (painel e XLSX).
            # "Nome do anúncio" e mantido apenas como fallback de nome e de chave,
            # para exports por conjunto/anuncio que nao trazem o nome da campanha.
            dfs_meta = []
            faltou_nome_campanha = False
            for f in files_meta:
                df_tmp = pd.read_csv(f, sep=',', encoding='utf-8-sig')
                if 'Valor gasto (BRL)' in df_tmp.columns:
                    df_tmp = df_tmp.rename(columns={'Valor gasto (BRL)': 'Valor usado (BRL)'})

                tem_camp = 'Nome da campanha' in df_tmp.columns
                tem_anun = 'Nome do anúncio' in df_tmp.columns
                if not tem_camp and not tem_anun:
                    raise ValueError(
                        f"O arquivo '{getattr(f, 'name', 'Meta')}' não tem nem a coluna "
                        f"'Nome da campanha' nem 'Nome do anúncio'."
                    )
                if not tem_camp:
                    faltou_nome_campanha = True
                    df_tmp['Nome da campanha'] = df_tmp['Nome do anúncio']
                if not tem_anun:
                    df_tmp['Nome do anúncio'] = df_tmp['Nome da campanha']

                dfs_meta.append(df_tmp[['Nome da campanha', 'Nome do anúncio',
                                        'Valor usado (BRL)', 'Identificação da campanha']])

            df_m = pd.concat(dfs_meta, ignore_index=True)

            if faltou_nome_campanha:
                st.info(
                    "ℹ️ Pelo menos um arquivo do Meta não tem a coluna **Nome da campanha** — "
                    "nessas linhas o painel exibe o nome do anúncio. Para ver o nome da campanha, "
                    "adicione essa coluna na exportação do Gerenciador de Anúncios."
                )

            if len(files_meta) > 1:
                st.info(f"📂 {len(files_meta)} arquivos do Meta consolidados — {len(df_m)} linhas no total.")

            df_a = pd.read_csv(file_adx, sep=';', encoding='utf-8-sig')

            # ── CHAVES DE PAREAMENTO ──────────────────────────────────────
            # A chave sai do nome da campanha; se ele nao render um numero
            # utilizavel, cai para o nome do anuncio.
            def chave_meta(row):
                k = gerar_chave(row['Nome da campanha'])
                if tem_numero(k):
                    return k
                k2 = gerar_chave(row['Nome do anúncio'])
                return k2 if tem_numero(k2) else k

            df_m['core'] = df_m.apply(chave_meta, axis=1)
            df_a['core'] = df_a['utm_campaign'].map(gerar_chave)

            # ── AUDITORIA: numeros de campanha repetidos no mesmo arquivo ─
            # Como a chave e apenas o numero, duas campanhas diferentes com o
            # mesmo numero (ex.: 164 para ARG e 164 para MEX no mesmo dia)
            # seriam somadas na mesma linha. Avisa se isso acontecer.
            dup = (df_m.groupby('core')['Nome da campanha']
                   .nunique().reset_index(name='n'))
            dup = dup[dup['n'] > 1]
            if not dup.empty:
                nomes_dup = (df_m[df_m['core'].isin(dup['core'])]
                             .sort_values('core')[['core', 'Nome da campanha', 'Nome do anúncio']])
                st.warning(
                    "⚠️ Números de campanha repetidos no Meta — estas linhas "
                    "serão somadas na mesma chave de pareamento:"
                )
                st.dataframe(nomes_dup, use_container_width=True, hide_index=True)

            # Nome da campanha, para exibir no painel/XLSX em vez da chave crua.
            # Vários anúncios da mesma campanha compartilham o mesmo nome.
            nome_exibicao = df_m.groupby('core')['Nome da campanha'].first().to_dict()

            # Converte receitas do AdX para float
            df_a['G_USD'] = (df_a['Ganhos'].astype(str)
                             .str.replace('$', '', regex=False)
                             .str.replace(',', '', regex=False)
                             .astype(float))

            # Linhas numéricas (ID da campanha) → chave via mapa de IDs do Meta
            df_m['ID_str'] = df_m['Identificação da campanha'].astype(str).str.strip()
            id_para_core = df_m.set_index('ID_str')['core'].to_dict()
            mask_numerica = df_a['utm_campaign'].astype(str).str.match(r'^\d{7,}$', na=False)
            df_a.loc[mask_numerica, 'core'] = (df_a.loc[mask_numerica, 'utm_campaign']
                                               .astype(str).str.strip().map(id_para_core))

            # ── AUDITORIA: receita do AdX sem campanha correspondente ─────
            chaves_meta = set(df_m['core'])
            orfas = df_a[~df_a['core'].isin(chaves_meta)].copy()
            orfas = orfas[orfas['G_USD'] > 0]
            if not orfas.empty:
                st.warning(
                    f"⚠️ {len(orfas)} linha(s) do AdX somando "
                    f"$ {orfas['G_USD'].sum():,.2f} não casaram com nenhuma campanha do Meta."
                )
                st.dataframe(orfas[['utm_campaign', 'Visitantes', 'G_USD']],
                             use_container_width=True, hide_index=True)

            # ── AGREGAÇÃO E JUNÇÃO ───────────────────────────────────────
            adx_g = df_a.dropna(subset=['core']).groupby('core')['G_USD'].sum().reset_index()
            meta_g = df_m.groupby('core')['Valor usado (BRL)'].sum().reset_index()
            merged = pd.merge(meta_g, adx_g, on='core', how='outer').fillna(0)
            merged['core'] = merged['core'].map(lambda k: nome_exibicao.get(k, k))
            merged.columns = ['Campanha', 'Investimento', 'Receita (USD)']

            # ── CONSOLIDACAO DA RECEITA RESIDUAL ──────────────────────────
            # Linhas com investimento 0 sao receita do AdX sem custo do dia:
            # visitas soltas de campanhas antigas ja pausadas, trafego organico
            # residual, lixo de tracking. Uma a uma, poluem o painel. Em vez de
            # descarta-las (o que faria o total divergir do AdX), todas viram
            # UMA linha agregada, fixada no fim da tabela.
            sem_inv = merged[merged['Investimento'] <= 0].copy()
            merged = merged[merged['Investimento'] > 0].copy()

            n_residuais = len(sem_inv)
            usd_residual = float(sem_inv['Receita (USD)'].sum())

            if usd_residual > 0:
                merged = pd.concat([
                    merged,
                    pd.DataFrame([{
                        'Campanha': LINHA_RESIDUAL,
                        'Investimento': 0.0,
                        'Receita (USD)': usd_residual
                    }])
                ], ignore_index=True)
                st.caption(
                    f"ℹ️ {n_residuais} linha(s) sem investimento foram agrupadas em "
                    f"**{LINHA_RESIDUAL}** (R$ {usd_residual * cambio:,.2f}). "
                    f"Elas entram na receita total para bater com o AdX, mas não têm ROI."
                )
                with st.expander(f"🔎 Detalhe das {n_residuais} linha(s) residuais"):
                    st.dataframe(
                        sem_inv.assign(**{'Receita (BRL)': sem_inv['Receita (USD)'] * cambio})
                        [['Campanha', 'Receita (USD)', 'Receita (BRL)']]
                        .sort_values('Receita (USD)', ascending=False)
                        .style.format({'Receita (USD)': '$ {:,.2f}', 'Receita (BRL)': 'R$ {:,.2f}'}),
                        use_container_width=True, hide_index=True
                    )

            merged['Receita (BRL)'] = merged['Receita (USD)'] * cambio
            merged['Lucro'] = merged['Receita (BRL)'] - merged['Investimento']
            merged['ROI'] = merged.apply(
                lambda r: r['Lucro'] / r['Investimento'] if r['Investimento'] > 0 else 0, axis=1
            )

            # Totais gerais: incluem a receita residual, para reconciliar com o AdX.
            inv_t = merged['Investimento'].sum()
            rec_t = merged['Receita (BRL)'].sum()
            luc_t = rec_t - inv_t
            roi_t = luc_t / inv_t if inv_t > 0 else 0

            # Totais so das campanhas com investimento: e este o ROI analitico,
            # sem a receita residual inflando o indicador.
            com_inv = merged[merged['Investimento'] > 0]
            rec_ci = com_inv['Receita (BRL)'].sum()
            luc_ci = rec_ci - inv_t
            roi_ci = luc_ci / inv_t if inv_t > 0 else 0

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Investimento Total", f"R$ {inv_t:,.2f}")
            m2.metric("Receita Total", f"R$ {rec_t:,.2f}")
            m3.metric("Lucro Líquido", f"R$ {luc_t:,.2f}")
            m4.metric("ROI Geral", f"{roi_t:.2%}")

            if usd_residual > 0:
                st.caption(
                    f"Somente campanhas com investimento: receita R$ {rec_ci:,.2f} · "
                    f"lucro R$ {luc_ci:,.2f} · **ROI {roi_ci:.2%}**"
                )

            # Ordena por ROI, mas mantem a linha residual sempre no fim.
            merged_sorted = (merged
                             .assign(_fim=(merged['Campanha'] == LINHA_RESIDUAL).astype(int))
                             .sort_values(['_fim', 'ROI'], ascending=[True, False])
                             .drop(columns='_fim'))

            def estilo_residual(row):
                """Linha agregada de receita residual: cinza e neutra, sem cor de ROI."""
                if row['Campanha'] == LINHA_RESIDUAL:
                    return ['color: #8a8a8a; font-style: italic'] * len(row)
                return [''] * len(row)

            st.dataframe(merged_sorted.style.format({
                'Investimento': 'R$ {:,.2f}',
                'Receita (USD)': '$ {:,.2f}',
                'Receita (BRL)': 'R$ {:,.2f}',
                'Lucro': 'R$ {:,.2f}',
                'ROI': lambda v: '—' if pd.isna(v) else f'{v:.2%}'
            }).map(color_negative_red, subset=['Lucro', 'ROI'])
                .apply(estilo_residual, axis=1),
                use_container_width=True, hide_index=True)

            # ── EXPORTAÇÃO XLSX FORMATADO ─────────────────────────────────────
            def gerar_xlsx(df, data_ref, inv_total, usd_total, rec_total, luc_total, roi_total):
                import io
                from openpyxl import Workbook
                from openpyxl.styles import (PatternFill, Font, Alignment,
                                             Border, Side, GradientFill)
                from openpyxl.utils import get_column_letter

                wb = Workbook()
                ws = wb.active
                ws.title = "Processamento Diário"

                # Cores
                COR_HEADER_BG  = "1A3A5C"
                COR_HEADER_FG  = "FFFFFF"
                COR_VERDE      = "1A7A4A"
                COR_VERMELHO   = "C0392B"
                COR_VERDE_BG   = "E9F7EF"
                COR_VERM_BG    = "FDEDEC"
                COR_TOTAL_BG   = "1F4E79"
                COR_ALT        = "F4F6F8"

                thin = Side(style='thin', color="CCCCCC")
                borda = Border(left=thin, right=thin, top=thin, bottom=thin)

                # ── Título ──
                ws.merge_cells("A1:G1")
                ws["A1"] = f"ROI Intelligence System — {data_ref}"
                ws["A1"].font = Font(bold=True, size=14, color=COR_HEADER_FG)
                ws["A1"].fill = PatternFill("solid", fgColor=COR_HEADER_BG)
                ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 28

                # ── Subtítulo com totais ──
                ws.merge_cells("A2:G2")
                ws["A2"] = (f"Investimento: R$ {inv_total:,.2f}   |   "
                            f"Receita: R$ {rec_total:,.2f}   |   "
                            f"Lucro: R$ {luc_total:,.2f}   |   "
                            f"ROI Geral: {roi_total:.2%}")
                ws["A2"].font = Font(bold=True, size=12, color="FFFFFF")
                ws["A2"].fill = PatternFill("solid", fgColor="1F4E79")
                ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[2].height = 20

                # ── Cabeçalho ──
                headers = ["Campanha", "Investimento (R$)", "Receita (USD)",
                           "Receita (BRL)", "Lucro (R$)", "ROI"]
                colunas_df = ["Campanha", "Investimento", "Receita (USD)",
                              "Receita (BRL)", "Lucro", "ROI"]

                for col_idx, h in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col_idx, value=h)
                    cell.font = Font(bold=True, color=COR_HEADER_FG, size=12)
                    cell.fill = PatternFill("solid", fgColor="2E75B6")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = borda
                ws.row_dimensions[3].height = 18

                # ── Dados ──
                COR_RESIDUAL_BG = "EFEFEF"
                COR_RESIDUAL_FG = "6B6B6B"

                for row_idx, (_, row) in enumerate(df.iterrows(), 4):
                    is_alt = (row_idx % 2 == 0)
                    lucro_val = row["Lucro"]
                    roi_val   = row["ROI"]
                    is_residual = (str(row["Campanha"]).upper() == LINHA_RESIDUAL)

                    # Linha agregada de receita residual: cinza, neutra, sem ROI.
                    if is_residual:
                        vals = [str(row["Campanha"]).upper(), None, float(row["Receita (USD)"]),
                                float(row["Receita (BRL)"]), float(lucro_val), None]
                        fmts = [None, None, '"$"#,##0.00', 'R$ #,##0.00', 'R$ #,##0.00', None]
                        for col_idx, (v, fmt_str) in enumerate(zip(vals, fmts), 1):
                            cell = ws.cell(row=row_idx, column=col_idx,
                                           value=v if v is not None else "—")
                            cell.border = borda
                            cell.alignment = Alignment(horizontal="right" if col_idx > 1 else "left",
                                                       vertical="center")
                            cell.font = Font(size=12, italic=True, color=COR_RESIDUAL_FG,
                                             bold=(col_idx == 1))
                            cell.fill = PatternFill("solid", fgColor=COR_RESIDUAL_BG)
                            if fmt_str:
                                cell.number_format = fmt_str
                        ws.row_dimensions[row_idx].height = 16
                        continue

                    for col_idx, col_name in enumerate(colunas_df, 1):
                        val = row[col_name]
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = borda
                        cell.alignment = Alignment(horizontal="right" if col_idx > 1 else "left",
                                                   vertical="center")

                        if col_name == "Campanha":
                            cell.value = str(val).upper()
                            cell.font = Font(size=12, bold=True)
                            cell.fill = PatternFill("solid", fgColor=COR_ALT if is_alt else "FFFFFF")

                        elif col_name == "ROI":
                            cell.value = roi_val
                            cell.number_format = '0.00%'
                            is_neg = roi_val < 0
                            cell.font = Font(bold=True, color=COR_VERMELHO if is_neg else COR_VERDE, size=12)
                            cell.fill = PatternFill("solid", fgColor=COR_VERM_BG if is_neg else COR_VERDE_BG)

                        elif col_name == "Lucro":
                            cell.value = lucro_val
                            cell.number_format = 'R$ #,##0.00'
                            is_neg = lucro_val < 0
                            cell.font = Font(bold=True, color=COR_VERMELHO if is_neg else COR_VERDE, size=12)
                            cell.fill = PatternFill("solid", fgColor=COR_VERM_BG if is_neg else COR_VERDE_BG)

                        elif col_name == "Receita (USD)":
                            cell.value = float(val)
                            cell.number_format = '"$"#,##0.00'
                            cell.font = Font(size=12)
                            cell.fill = PatternFill("solid", fgColor=COR_ALT if is_alt else "FFFFFF")

                        else:  # Investimento, Receita (BRL)
                            cell.value = float(val)
                            cell.number_format = 'R$ #,##0.00'
                            cell.font = Font(size=12)
                            cell.fill = PatternFill("solid", fgColor=COR_ALT if is_alt else "FFFFFF")

                    ws.row_dimensions[row_idx].height = 16

                # ── Linha de totais ──
                tot_row = ws.max_row + 1
                totais_vals = ["TOTAL", inv_total, usd_total, rec_total, luc_total, roi_total]
                totais_fmt  = [None, 'R$ #,##0.00', '"$"#,##0.00', 'R$ #,##0.00', 'R$ #,##0.00', '0.00%']

                for col_idx, (val, fmt_str) in enumerate(zip(totais_vals, totais_fmt), 1):
                    cell = ws.cell(row=tot_row, column=col_idx, value=val)
                    cell.font = Font(bold=True, color="FFFFFF", size=12)
                    cell.fill = PatternFill("solid", fgColor=COR_TOTAL_BG)
                    cell.alignment = Alignment(horizontal="right" if col_idx > 1 else "left",
                                               vertical="center")
                    cell.border = borda
                    if fmt_str:
                        cell.number_format = fmt_str
                ws.row_dimensions[tot_row].height = 18

                # ── Larguras das colunas ──
                larguras = [38, 18, 15, 16, 16, 12]
                for i, w in enumerate(larguras, 1):
                    ws.column_dimensions[get_column_letter(i)].width = w

                # Salva em buffer
                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                return buf.getvalue()

            xlsx_bytes = gerar_xlsx(
                merged_sorted,
                data_referencia.strftime('%d/%m/%Y'),
                inv_t, merged_sorted['Receita (USD)'].sum(), rec_t, luc_t, roi_t
            )

            st.download_button(
                label="⬇️ Exportar XLSX",
                data=xlsx_bytes,
                file_name=f"roi_{data_referencia.strftime('%Y-%m-%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # Prepara as novas linhas para salvar (usadas tanto no save normal quanto na atualização)
            def build_new_rows(df, data_str):
                return [
                    [
                        data_str,
                        r['Campanha'].upper(),
                        round(float(r['Investimento']), 2),
                        round(float(r['Receita (USD)']), 2),
                        round(float(r['Receita (BRL)']), 2),
                        round(float(r['Lucro']), 2)
                    ]
                    for _, r in df.iterrows()
                ]

            if st.button("💾 Salvar no Google Sheets"):
                data_str = data_referencia.strftime('%Y-%m-%d')
                client = get_gspread_client()
                sheet = client.open_by_key(st.secrets["spreadsheet"]["id"]).worksheet("Historico")

                raw_existing = sheet.get_all_values()
                datas_existentes = set()
                if len(raw_existing) > 1:
                    for row in raw_existing[1:]:
                        if row:
                            datas_existentes.add(str(row[0]).strip()[:10])

                if data_str in datas_existentes:
                    # Armazena contexto para o diálogo de confirmação
                    st.session_state['aguardando_confirmacao'] = True
                    st.session_state['data_str_pendente'] = data_str
                    st.session_state['merged_pendente'] = merged.copy()
                else:
                    # Não há duplicata: salva diretamente
                    new_rows = build_new_rows(merged, data_str)
                    sheet.append_rows(new_rows, value_input_option='RAW')
                    st.session_state['aguardando_confirmacao'] = False
                    st.success(f"✅ {len(new_rows)} campanhas salvas para {data_str}!")

            # Diálogo de confirmação (persiste entre re-execuções via session_state)
            if st.session_state.get('aguardando_confirmacao'):
                data_str = st.session_state['data_str_pendente']
                st.warning(
                    f"⚠️ Já existem dados salvos para **{data_str}**. "
                    f"Deseja substituí-los pelos novos dados carregados?"
                )
                col_sim, col_nao, _ = st.columns([1, 1, 6])

                with col_sim:
                    if st.button("✅ Sim, atualizar", type="primary"):
                        try:
                            client2 = get_gspread_client()
                            sheet2 = client2.open_by_key(st.secrets["spreadsheet"]["id"]).worksheet("Historico")
                            all_rows = sheet2.get_all_values()
                            headers = all_rows[0]

                            # Identifica os índices das linhas da data a substituir (1-based no Sheets)
                            linhas_para_deletar = [
                                i + 2  # +1 pelo header, +1 porque Sheets é 1-based
                                for i, row in enumerate(all_rows[1:])
                                if row and str(row[0]).strip()[:10] == data_str
                            ]

                            # Deleta de baixo para cima para não deslocar índices
                            for idx in sorted(linhas_para_deletar, reverse=True):
                                sheet2.delete_rows(idx)

                            # Insere as novas linhas no final
                            merged_p = st.session_state['merged_pendente']
                            new_rows = build_new_rows(merged_p, data_str)
                            sheet2.append_rows(new_rows, value_input_option='RAW')

                            st.session_state['aguardando_confirmacao'] = False
                            st.success(f"✅ Dados de {data_str} atualizados com sucesso! ({len(new_rows)} campanhas)")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erro ao atualizar: {e}")

                with col_nao:
                    if st.button("❌ Não, cancelar"):
                        st.session_state['aguardando_confirmacao'] = False
                        st.info("Operação cancelada. Os dados existentes foram mantidos.")
                        st.rerun()

        except Exception as e:
            st.error(f"Erro: {e}")

with tab2:
    st.subheader("🔍 Análise de Desempenho Histórico")

    hoje = datetime.now().date()
    primeiro_dia_mes = hoje.replace(day=1)
    ultimo_mes_fim = primeiro_dia_mes - timedelta(days=1)
    ultimo_mes_inicio = ultimo_mes_fim.replace(day=1)

    opcao_data = st.selectbox("Selecione o Período", [
        "Hoje", "Ontem", "Últimos 7 dias", "Últimos 15 dias",
        "Mês Atual", "Mês Passado", "Personalizado"
    ])

    if opcao_data == "Hoje":
        start, end = hoje, hoje
    elif opcao_data == "Ontem":
        start = end = hoje - timedelta(days=1)
    elif opcao_data == "Últimos 7 dias":
        start, end = hoje - timedelta(days=6), hoje
    elif opcao_data == "Últimos 15 dias":
        start, end = hoje - timedelta(days=14), hoje
    elif opcao_data == "Mês Atual":
        start, end = primeiro_dia_mes, hoje
    elif opcao_data == "Mês Passado":
        start, end = ultimo_mes_inicio, ultimo_mes_fim
    else:
        c1, c2 = st.columns(2)
        start = c1.date_input("Início", hoje - timedelta(days=30))
        end = c2.date_input("Fim", hoje)

    if st.button("Consultar Histórico"):
        try:
            client = get_gspread_client()
            sheet = client.open_by_key(st.secrets["spreadsheet"]["id"]).worksheet("Historico")
            raw = sheet.get_all_values()

            if len(raw) < 2:
                st.warning("Sem dados na planilha.")
            else:
                headers = raw[0]
                df_h = pd.DataFrame(raw[1:], columns=headers)

                for col in ['Investimento', 'Receita (USD)', 'Receita (BRL)', 'Lucro']:
                    if col in df_h.columns:
                        df_h[col] = df_h[col].apply(clean_numeric)

                df_h['ROI'] = df_h.apply(
                    lambda r: r['Lucro'] / r['Investimento'] if r['Investimento'] > 0 else 0, axis=1
                )
                df_h['Data_Ref'] = pd.to_datetime(df_h['Data_Ref'], errors='coerce').dt.date

                df_final = df_h[(df_h['Data_Ref'] >= start) & (df_h['Data_Ref'] <= end)].copy()

                if df_final.empty:
                    st.warning("Sem dados para este período.")
                else:
                    # ── CARDS ──────────────────────────────────────────────────
                    inv_h = df_final['Investimento'].sum()
                    rec_h = df_final['Receita (BRL)'].sum()
                    luc_h = rec_h - inv_h
                    roi_h = luc_h / inv_h if inv_h > 0 else 0

                    h1, h2, h3, h4 = st.columns(4)
                    h1.metric("💰 Investimento", f"R$ {inv_h:,.2f}")
                    h2.metric("📥 Receita", f"R$ {rec_h:,.2f}")
                    h3.metric("💵 Lucro", f"R$ {luc_h:,.2f}")
                    h4.metric("📊 ROI Médio", f"{roi_h:.2%}")

                    st.divider()

                    # ── TABELAS ────────────────────────────────────────────────
                    st.markdown("### 📋 Tabelas")
                    tb_resumo, tb_detalhado, tb_diario = st.tabs([
                        "Resumo por Campanha", "Detalhado por Dia", "📅 Consolidado por Dia"
                    ])

                    fmt = {
                        'Investimento': 'R$ {:,.2f}',
                        'Receita (USD)': '$ {:,.2f}',
                        'Receita (BRL)': 'R$ {:,.2f}',
                        'Lucro': 'R$ {:,.2f}',
                        'ROI': '{:.2%}'
                    }

                    with tb_resumo:
                        resumo = df_final.groupby('Campanha').agg(
                            Investimento=('Investimento', 'sum'),
                            Receita_USD=('Receita (USD)', 'sum'),
                            Receita_BRL=('Receita (BRL)', 'sum'),
                            Lucro=('Lucro', 'sum'),
                            Dias=('Data_Ref', 'nunique')
                        ).reset_index()
                        resumo['ROI'] = resumo.apply(
                            lambda r: r['Lucro'] / r['Investimento'] if r['Investimento'] > 0 else 0, axis=1
                        )
                        resumo = resumo.rename(columns={
                            'Receita_USD': 'Receita (USD)',
                            'Receita_BRL': 'Receita (BRL)'
                        })
                        resumo = resumo.sort_values('ROI', ascending=False)

                        st.dataframe(
                            resumo[['Campanha', 'Dias', 'Investimento', 'Receita (USD)', 'Receita (BRL)', 'Lucro', 'ROI']]
                            .style.format({**fmt, 'Dias': '{:.0f}'})
                            .map(color_negative_red, subset=['Lucro', 'ROI']),
                            use_container_width=True, hide_index=True
                        )

                    with tb_detalhado:
                        det = df_final[['Data_Ref', 'Campanha', 'Investimento', 'Receita (USD)', 'Receita (BRL)', 'Lucro', 'ROI']].copy()
                        det = det.sort_values(['Data_Ref', 'Campanha'])
                        st.dataframe(
                            det.style.format(fmt)
                            .map(color_negative_red, subset=['Lucro', 'ROI']),
                            use_container_width=True, hide_index=True
                        )

                    with tb_diario:
                        # Agrupa tudo por data — uma linha por dia
                        cons = df_final.groupby('Data_Ref').agg(
                            Investimento=('Investimento', 'sum'),
                            Receita_USD=('Receita (USD)', 'sum'),
                            Receita_BRL=('Receita (BRL)', 'sum'),
                            Lucro=('Lucro', 'sum')
                        ).reset_index()
                        cons['ROI'] = cons.apply(
                            lambda r: r['Lucro'] / r['Investimento'] if r['Investimento'] > 0 else 0, axis=1
                        )
                        cons = cons.rename(columns={
                            'Receita_USD': 'Receita (USD)',
                            'Receita_BRL': 'Receita (BRL)'
                        })
                        cons = cons.sort_values('Data_Ref')

                        # Linha de totais do período
                        total_inv  = cons['Investimento'].sum()
                        total_usd  = cons['Receita (USD)'].sum()
                        total_brl  = cons['Receita (BRL)'].sum()
                        total_luc  = cons['Lucro'].sum()
                        total_roi  = total_luc / total_inv if total_inv > 0 else 0

                        totais = pd.DataFrame([{
                            'Data_Ref': 'TOTAL',
                            'Investimento': total_inv,
                            'Receita (USD)': total_usd,
                            'Receita (BRL)': total_brl,
                            'Lucro': total_luc,
                            'ROI': total_roi
                        }])

                        cons_display = pd.concat([cons, totais], ignore_index=True)
                        cons_display['Data_Ref'] = cons_display['Data_Ref'].astype(str)

                        def highlight_total(row):
                            if row['Data_Ref'] == 'TOTAL':
                                return ['font-weight: bold; background-color: #1e1e2e; color: white'] * len(row)
                            return [''] * len(row)

                        st.dataframe(
                            cons_display[['Data_Ref', 'Investimento', 'Receita (USD)', 'Receita (BRL)', 'Lucro', 'ROI']]
                            .style
                            .format({**fmt, 'Data_Ref': '{}'})
                            .map(color_negative_red, subset=['Lucro', 'ROI'])
                            .apply(highlight_total, axis=1),
                            use_container_width=True, hide_index=True
                        )

                        # Botão de download CSV
                        csv_data = cons[['Data_Ref', 'Investimento', 'Receita (USD)', 'Receita (BRL)', 'Lucro', 'ROI']].copy()
                        csv_data['Data_Ref'] = csv_data['Data_Ref'].astype(str)
                        csv_data['ROI'] = (csv_data['ROI'] * 100).round(2).astype(str) + '%'
                        st.download_button(
                            label="⬇️ Baixar CSV",
                            data=csv_data.to_csv(index=False, sep=';', decimal=',').encode('utf-8'),
                            file_name=f"consolidado_{start}_{end}.csv",
                            mime='text/csv'
                        )

                    st.divider()

                    # ── GRÁFICOS ───────────────────────────────────────────────
                    st.markdown("### 📈 Gráficos")

                    # 1. Evolução diária — ROI e Lucro
                    st.markdown("#### Evolução Diária — ROI e Lucro")
                    diario = df_final.groupby('Data_Ref').agg(
                        Investimento=('Investimento', 'sum'),
                        Receita_BRL=('Receita (BRL)', 'sum'),
                        Lucro=('Lucro', 'sum')
                    ).reset_index()
                    diario['ROI (%)'] = diario.apply(
                        lambda r: (r['Lucro'] / r['Investimento'] * 100) if r['Investimento'] > 0 else 0, axis=1
                    )
                    diario['Data_Ref'] = diario['Data_Ref'].astype(str)

                    fig1 = go.Figure()
                    fig1.add_trace(go.Bar(
                        x=diario['Data_Ref'], y=diario['Lucro'],
                        name='Lucro (R$)',
                        marker_color=['#2ecc71' if v >= 0 else '#e74c3c' for v in diario['Lucro']],
                        yaxis='y1'
                    ))
                    fig1.add_trace(go.Scatter(
                        x=diario['Data_Ref'], y=diario['ROI (%)'],
                        name='ROI (%)', mode='lines+markers',
                        line=dict(color='#3498db', width=2),
                        marker=dict(size=7), yaxis='y2'
                    ))
                    fig1.update_layout(
                        yaxis=dict(title='Lucro (R$)', side='left'),
                        yaxis2=dict(title='ROI (%)', side='right', overlaying='y', ticksuffix='%'),
                        legend=dict(orientation='h', y=1.1),
                        hovermode='x unified', height=380
                    )
                    st.plotly_chart(fig1, use_container_width=True)

                    # 2. Investimento vs Receita por dia
                    st.markdown("#### Investimento vs Receita por Dia")
                    fig2 = go.Figure()
                    fig2.add_trace(go.Bar(
                        x=diario['Data_Ref'], y=diario['Investimento'],
                        name='Investimento', marker_color='#e67e22'
                    ))
                    fig2.add_trace(go.Bar(
                        x=diario['Data_Ref'], y=diario['Receita_BRL'],
                        name='Receita (BRL)', marker_color='#2ecc71'
                    ))
                    fig2.update_layout(
                        barmode='group', yaxis_title='R$',
                        legend=dict(orientation='h', y=1.1),
                        hovermode='x unified', height=350
                    )
                    st.plotly_chart(fig2, use_container_width=True)

                    # 3. Ranking de campanhas por ROI
                    st.markdown("#### Ranking de Campanhas por ROI no Período")
                    ranking = resumo.sort_values('ROI')
                    ranking = ranking.copy()
                    ranking['ROI (%)'] = ranking['ROI'] * 100
                    ranking['cor'] = ranking['ROI (%)'].apply(lambda v: '#2ecc71' if v >= 0 else '#e74c3c')

                    fig3 = go.Figure(go.Bar(
                        x=ranking['ROI (%)'],
                        y=ranking['Campanha'],
                        orientation='h',
                        marker_color=ranking['cor'].tolist(),
                        text=ranking['ROI (%)'].apply(lambda v: f'{v:.1f}%'),
                        textposition='outside'
                    ))
                    fig3.update_layout(
                        xaxis=dict(title='ROI (%)', ticksuffix='%'),
                        height=max(350, len(ranking) * 38),
                        margin=dict(l=20, r=80)
                    )
                    st.plotly_chart(fig3, use_container_width=True)

                    # 4. ROI por campanha ao longo do tempo
                    st.markdown("#### ROI por Campanha ao Longo do Tempo")
                    camp_diario = df_final.copy()
                    camp_diario['Data_Ref'] = camp_diario['Data_Ref'].astype(str)
                    camp_diario['ROI (%)'] = camp_diario['ROI'] * 100

                    fig4 = px.line(
                        camp_diario.sort_values('Data_Ref'),
                        x='Data_Ref', y='ROI (%)',
                        color='Campanha', markers=True,
                        labels={'Data_Ref': 'Data', 'ROI (%)': 'ROI (%)'}
                    )
                    fig4.add_hline(y=0, line_dash='dash', line_color='gray', opacity=0.5)
                    fig4.update_layout(
                        yaxis_ticksuffix='%',
                        hovermode='x unified',
                        legend=dict(orientation='h', y=-0.3),
                        height=450
                    )
                    st.plotly_chart(fig4, use_container_width=True)

        except Exception as e:
            st.error(f"Erro ao consultar: {e}")
